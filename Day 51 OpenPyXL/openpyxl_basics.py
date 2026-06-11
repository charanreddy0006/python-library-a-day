from openpyxl import Workbook, load_workbook

# Create workbook
workbook = Workbook()

sheet = workbook.active

sheet.title = "Students"

# Add data
sheet["A1"] = "Name"
sheet["B1"] = "Marks"

sheet["A2"] = "Chakri"
sheet["B2"] = 95

sheet["A3"] = "Sai "
sheet["B3"] = 88

# Save workbook
workbook.save("students.xlsx")

print("Excel file created successfully!")

# Read workbook
wb = load_workbook("students.xlsx")

sheet = wb["Students"]

print("\nStudent Data:")

for row in sheet.iter_rows(values_only=True):
    print(row)