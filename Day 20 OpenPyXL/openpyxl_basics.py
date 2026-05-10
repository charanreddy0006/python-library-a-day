from openpyxl import Workbook, load_workbook

# --- create workbook ---
workbook = Workbook()

# --- active sheet ---
sheet = workbook.active

sheet.title = "Student Data"

# --- write data ---
sheet["A1"] = "Name"
sheet["B1"] = "Marks"

sheet["A2"] = "Chakri"
sheet["B2"] = 95

sheet["A3"] = "Rahul"
sheet["B3"] = 88

# --- save workbook ---
workbook.save("students.xlsx")

print("Excel file created successfully")

# --- load workbook ---
loaded_workbook = load_workbook("students.xlsx")

loaded_sheet = loaded_workbook["Student Data"]

# --- read values ---
print("\nReading Excel Data:\n")

for row in loaded_sheet.iter_rows(values_only=True):
    print(row)